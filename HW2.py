#!/usr/bin/python3

import openpyxl
from openpyxl import load_workbook

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
st_count = 0
st_score = list()
dic = dict()
i = 1
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column = 3).value * 0.3
        sum_v += ws.cell(row = row_id, column = 4).value * 0.35
        sum_v += ws.cell(row = row_id, column = 5).value * 0.34
        sum_v += ws.cell(row = row_id, column = 6).value
        ws.cell(row = row_id, column = 7).value = sum_v
        st_score.append(sum_v)
        st_count += 1
        dic[i] = sum_v
    row_id += 1
    i += 1

result_dict = dict(sorted(dic.items(), reverse = True, key = lambda x:x[1]))
st_score.sort(reverse = True)
dict_keys = list(result_dict.keys())

A_list = list()
B_list = list()
C_list = list()
st_A = list()
st_B = list()
st_C = list()
tmp = 0

while (tmp < st_count - 2):
    if st_score.count(st_score[tmp]) < 2:
        if tmp + 1 <= st_count * 0.3:
            A_list.append(st_score[tmp])
            st_A.append(dict_keys[tmp])
            tmp += 1
        elif tmp + 1 <= st_count * 0.7:
            B_list.append(st_score[tmp])
            st_B.append(dict_keys[tmp])
            tmp += 1
        else:
            C_list.append(st_score[tmp])
            st_C.append(dict_keys[tmp])
            tmp += 1

    else:
        n = st_score.count(st_score[tmp])
        if tmp + n <= st_count * 0.3:
            for j in range(n):
                A_list.append(tmp + j)
                st_A.append(dict_keys[tmp + j])
            tmp += n
        elif tmp + n <= st_count * 0.7:
            for j in range(n):
                B_list.append(tmp + j)
                st_B.append(dict_keys[tmp + j])
            tmp += n
        else:
            for j in range(n):
                C_list.append(tmp + j)
                st_C.append(dict_keys[tmp + j])
            tmp += n

tmp = 0
i = 0
while(tmp < len(A_list)):
    if st_score.count(st_score[tmp]) > 1:
        n = st_score.count(st_score[tmp])
        if i + n <= len(A_list) * 0.5:
            for j in range(n):
                ws.cell(row = st_A[i + j], column = 8, value = "A+")
        else:
            for j in range(n):
                ws.cell(row = st_A[i + j], column = 8, value = "A0")
        tmp += n
        i += n
    else:
        if tmp + 1 <= len(A_list) * 0.5:
            ws.cell(row = st_A[i], column = 8, value = "A+")
        else:
            ws.cell(row = st_A[i], column = 8, value = "A0")
        tmp += 1
        i += 1

i = 0
while (tmp < len(A_list) + len(B_list)):
    if st_score.count(st_score[tmp]) > 1:
        n = st_score.count(st_score[tmp])
        if i + n <= len(B_list) * 0.5:
            for j in range(n):
                ws.cell(row=st_B[i + j], column=8, value="B+")
        else:
            for j in range(n):

                ws.cell(row=st_B[i + j], column=8, value="B0")
        tmp += n
        i += n
    else:
        if tmp <= len(B_list) * 0.5:
            ws.cell(row=st_B[i], column=8, value="B+")
        else:
            ws.cell(row=st_B[i], column=8, value="B0")
        tmp += 1
        i += 1

i = 0
print(st_C)
while (tmp < st_count):
    if st_score.count(st_score[tmp]) > 1:
        n = st_score.count(st_score[tmp])
        if i + n <= len(C_list) * 0.5:
            for j in range(n):
                ws.cell(row=st_C[i + j], column=8, value="C+")
        else:
            for j in range(n):
                ws.cell(row=st_C[i + j], column=8, value="C0")
        tmp += n
        i += n
    else:
        if tmp + 1 <= len(C_list) * 0.5:
            ws.cell(row=st_C[i], column=8, value="C+")
        else:
            ws.cell(row=st_C[i], column=8, value="C0")
        tmp += 1
        i += 1

wb.save("student.xlsx")